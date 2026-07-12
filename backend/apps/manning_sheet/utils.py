import logging
logger = logging.getLogger(__name__)

import os
import re
import ast
import math
import json
import requests
import datetime
import numpy as np
import pandas as pd
import networkx as nx

from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from django.conf import settings
from openpyxl.styles import Font
from django.db import transaction
from rest_framework import status
from collections import defaultdict
from openpyxl.utils import get_column_letter
from django.db.models import Case, When, IntegerField, Sum, Value, Q

from config.utils import truncate_table
from apps.data_engine.models import AttendanceMaster
from .models import PushNotification, DDayData, EmployeesOnHold

# Variables for OptaFloor APIS
SKILL_MATRIX_API_URL = "https://optafloor-laguna.jaza-soft.com/v1/api/external/skill-matrix"
OPERATION_API_URL = "https://optafloor-laguna.jaza-soft.com/v1/api/external/operations"
STYLEOB_API_URL = "https://optafloor-laguna.jaza-soft.com/v1/api/external/ob-templates"
WIP_API_URL = "https://optafloor-laguna.jaza-soft.com/v1/api/external/wip"
PAGE_SIZE = 2000 # Max page size
API_KEY = "QBnzi5JBIwpUohMll8UTT75K.TEFHVU5BX0FJ"
DEPARTMENT = "Sewing"
VALID_LINES = {"Line 1", "Line 2", "Line 3", "Line 4", "Line 5", "Line 6", "Line 7", "Line 8", "Line 9A", "Line 9B"} # Define valid lines
FACTORY_FLOOR_MAPPING = {
    "Line 1": ("Factory 1", "Floor 1"),
    "Line 2": ("Factory 1", "Floor 1"),
    "Line 3": ("Factory 2", "Floor 1"),
    "Line 4": ("Factory 2", "Floor 1"),
    "Line 5": ("Factory 3", "Floor 2"),
    "Line 6": ("Factory 3", "Floor 2"),
    "Line 7": ("Factory 4", "Floor 2"),
    "Line 8": ("Factory 4", "Floor 2"),
    "Line 9": ("Factory 5", "Floor 2"),
    "Line 10": ("Factory 5", "Floor 2")
}




def convert_to_excel_data(data):
    excel_data = BytesIO()
    data = dict(data)
    
    # Create an Excel writer with BytesIO
    with pd.ExcelWriter(excel_data, engine="openpyxl") as writer:
        for section, section_data in data.items():
            # Extract general information
            general_info_df = pd.DataFrame([section_data["general_information"]])
            table_data_df = pd.DataFrame(section_data["table_data"])
            
            # Write General Information at the top
            general_info_df.to_excel(writer, index=False, sheet_name=section, startrow=1)
            start_row = len(general_info_df) + 3  # Leave space after general info
            
            # Write table data below general info
            table_data_df.to_excel(writer, index=False, sheet_name=section, startrow=start_row)
            
            # Auto-adjust column width based on header and data independently for each section
            worksheet = writer.sheets[section]
            
            # Adjust column widths for general_info_df
            for col_idx, col_name in enumerate(general_info_df.columns):
                max_length = max(len(str(col_name)), *[len(str(x)) for x in general_info_df[col_name].astype(str)])
                col_letter = get_column_letter(col_idx + 1)  # Convert column index to Excel letter
                worksheet.column_dimensions[col_letter].width = max_length + 2  # Adjust padding
            
            # Adjust column widths for table_data_df separately
            for col_idx, col_name in enumerate(table_data_df.columns):
                if col_name == "SMV":
                    continue  # Skip adjusting column width for "SMV"
                max_length = max(len(str(col_name)), *[len(str(x)) for x in table_data_df[col_name].astype(str)])
                col_letter = get_column_letter(col_idx + 1)  # Convert column index to Excel letter
                worksheet.column_dimensions[col_letter].width = max_length + 2  # Adjust padding

    # Reset pointer to the beginning
    excel_data.seek(0)
    
    return excel_data


def is_array_of_objects(data):
    return isinstance(data, list) and all(isinstance(item, dict) for item in data)


# Function to fetch data from different OptaFloor APIs
def fetch_data(all_data, page=0, api_url=None, department=None, transform_df=True, poRef=None, style=None, color=None, line=None):
    try:
        api_endpoint = f"{api_url}?size={PAGE_SIZE}&page={page}"
        
        if department:
            api_endpoint += f"&department={department}"
        
        if poRef and style and color:
            api_endpoint += f"&poRef={poRef}&style={style}&color={color}&line={line}"
        
        logger.info(f"Fetching API: {api_endpoint}")
        response = requests.get(api_endpoint, headers={'x-api-key': API_KEY})
        logger.info("Response: ", response)

        if response.status_code != 200:
            empty_df = pd.DataFrame()
            return empty_df

        response.raise_for_status()
        data = response.json()

        # Example
        if is_array_of_objects(data):
            return pd.DataFrame(data)
        else:
            if data and "content" in data:
                all_data.extend(data["content"])

            logger.info(f"Fetched page {page + 1} of {data.get('totalPages', 'Unknown')}")
            
            if not data.get("last", True):
                return fetch_data(all_data, page + 1, api_url, department, transform_df)  # ✅ Return result from recursive call


        if transform_df:
            # Converting the data into a dataframe for cleaning
            df = convert_to_dataframe(all_data)

            return df  # ✅ Ensure df is returned
        else:
            return all_data
    except Exception as e:
        logger.info(f"Error fetching data: {e}")



# To convert data from an array into a pandas dataframe
def convert_to_dataframe(all_data):
    try:
        if not all_data:
            logger.info("No data to save.")
            return False
        df = pd.DataFrame(all_data)
        all_data.clear()
        return df
    except Exception as e:
        logger.info(f"Error generating dataframe: {e}")
        return False



# To save dataframe into a csv file
def save_to_csv(file_name, df):
    try:
        file_path = f"{file_name}.csv"
        
        # Delete existing file if it exists
        if os.path.exists(file_path):
            os.remove(file_path)
            logger.info(f"Existing file {file_path} deleted.")
        
        # Save DataFrame to CSV
        df.to_csv(file_path, index=False, encoding='utf-8')
        logger.info(f"Data saved to {file_path}")
    except Exception as e:
        logger.info(f"Error saving CSV: {e}")



# To get filename from the API passed
def get_file_name(api_url):
    try:
        return api_url.split("external/")[1]
    except Exception as e:
        logger.info(f"Error in getting file name: {e}")
        return "output"



# To fetch dat from OptaFloor's skill-matrix api
def fetch_skill_matrix():
    try:
        all_data = []
        logger.info("Fetching Skill Matrix")
        df = fetch_data(all_data, 0, SKILL_MATRIX_API_URL)
        return df
    except Exception as e:
        logger.info(f"Error in Skill Matrix API: {e}")



# To fetch dat from OptaFloor's master operations api
def fetch_operations():
    try:
        all_data = []
        logger.info("Fetching Operations")
        df = fetch_data(all_data, 0, OPERATION_API_URL, DEPARTMENT)
        return df
    except Exception as e:
        logger.info(f"Error in Operations API: {e}")



# To fetch dat from OptaFloor's master operations api
def fetch_style_ob():
    try:
        all_data = []
        logger.info("Fetching Style OB Operations")
        raw_data = fetch_data(all_data, 0, STYLEOB_API_URL, transform_df=False)
        df = creates_dataframe(raw_data)
        return df
    except Exception as e:
        logger.info(f"Error in Style OB API: {e}")



# To fetch dat from OptaFloor's master operations api
def fetch_wip(poRef, style, color, line):
    try:
        all_data = []
        logger.info("Fetching WIP Data")
        raw_data = fetch_data(all_data, 0, WIP_API_URL, poRef=poRef, style=style, color=color, line=line, transform_df=False)
        return raw_data
    except Exception as e:
        logger.info(f"Error in WIP Data API: {e}")



# Custom function to transform style ob response into a dataframe
def creates_dataframe(raw_data):
    rows = []
    # Loop through the data and collect rows
    for item in raw_data:
        name = item.get('name')
        department = item.get('department')
        product_group = item.get('productGroup')
        created_at = item.get('createdAt')
        products = item.get('products', [None])

        for product in (products if products else [None]):
            for operation in item.get('operationList', []):
                operation_name = operation.get('name')
                operation_code = operation.get('code')
                operation_type = operation.get('type')
                section = operation.get('section')  # This can be None, that's fine
                machinist = operation.get('machinist')
                dependencies = ", ".join(operation.get('dependencies', [])) if operation.get('dependencies') else None

                rows.append({
                    "Name": name,
                    "Department": department,
                    "Product Group": product_group,
                    "Product": product,
                    "Created At": created_at,
                    "Operation Name": operation_name,
                    "Operation Code": operation_code,
                    "Operation Type": operation_type,
                    "Section": section,
                    "Machinist": machinist,
                    "Dependencies": dependencies
                })
    
    # Create a DataFrame from the list of rows
    df = pd.DataFrame(rows)
    return df



# Function to keep only the first valid element from 'lines'
def keep_first_valid_line(line_list):
    if isinstance(line_list, str):  # Convert string representation of list to actual list
        try:
            line_list = ast.literal_eval(line_list)  # Convert to list if it's a string
        except (ValueError, SyntaxError):
            return None  # If conversion fails, return None
    
    if isinstance(line_list, list):  # Ensure it's a list
        valid_only = [line for line in line_list if line in VALID_LINES]  # Filter valid lines
        return valid_only[0] if valid_only else None  # Keep only the first valid element
    return None



# To clean and merge the skill matrix and operation dataframe
def merge_dataframe(df_skill_matrix, df_operations):
    try:



        # Snippet for filtering the skill matrix dataframe using lastUpdatedAt and keep only the latest record
        # Convert to datetime
        # df_skill_matrix["lastUpdatedAt"] = pd.to_datetime(df_skill_matrix["lastUpdatedAt"], errors='coerce')
        # df_skill_matrix["lastUpdatedAt"] = df_skill_matrix["lastUpdatedAt"].dt.date

        # Get today's date
        # today = datetime.today().date()

        # Calculate six months ago
        # six_months_ago = today - relativedelta(months=6)

        # Filter
        # df_skill_matrix = df_skill_matrix[df_skill_matrix['lastUpdatedAt'] >= six_months_ago]
        # Keep only the latest record for each employeeId
        # df_skill_matrix = df_skill_matrix.sort_values(by='lastUpdatedAt', ascending=False)
        # df_skill_matrix = df_skill_matrix.drop_duplicates(subset='employeeId', keep='first')

        df_skill_matrix = process_skill_matrix_data(df_skill_matrix)
        
        df_merged = df_skill_matrix.merge(
            df_operations, 
            left_on="operationCode", 
            right_on="code", 
            how="left"
        )
        # Remove rows where 'employeeId' or 'code' is blank (empty string or NaN)
        df_merged.dropna(subset=['employeeId', 'code'], inplace=True)

        # Select relevant columns and rename properly
        df_merged = df_merged[[
            "operatorName", "employeeId", "lines", "type_x", "peakCapacity", 
            "peakPerformance", "avgCapacity", "avgPerformance", "lastUpdatedAt", 
            "section_x", "operationName", "operationCode", "machinist_x",
            "department", "section_y", "name", "code", "machinist_y", "type_y", "sam", "machineType"
        ]].rename(columns={
            "type_x": "skillType",
            "section_x": "section",
            "section_y": "operationSection",
            "machinist_x": "machinist",
            "machinist_y": "operationMachinist",
            "type_y": "operationType"
        })

        # df_merged['DESIGNATION'] = df_merged['machinist'].apply(lambda x: 'Machinist' if x else 'Non-Machinist')

        df_merged['STATUS'] = 'Critical'  # Add a new column with a constant value

        df_merged['AVERAGE CAPACITY'] = df_merged['avgCapacity'].fillna(0) * 9 # Multiply avgCapacity by 9
        df_merged['PEAK CAPACITY'] = df_merged['peakCapacity'].fillna(0) * 9 # Multiply peakCapacity by 9

        df_merged['AVERAGE PERFORMANCE'] = df_merged['avgPerformance'].fillna(0) * 9 # Multiply avgPerformance by 9
        df_merged['PEAK PERFORMANCE'] = df_merged['peakPerformance'].fillna(0) * 9 # Multiply peakPerformance by 9


        # # List of columns to drop
        # columns_to_drop = ['operationSection', 'lastUpdatedAt', 'operationName', 'operationCode', 'operationMachinist', 'operationType']

        # # # Drop the columns
        # df_merged = df_merged.drop(columns=columns_to_drop)

        df_merged = df_merged.rename(columns={
            'employeeId': 'EMPLOYEE ID',
            'operatorName' : 'EMPLOYEE NAME', 
            'lines': 'LINE',
            'section' : 'SECTION',
            'code': 'CODE',
            'name': 'OPERATION',
            'skillType': 'TYPE',
            'sam': 'SAM',
            'peakCapacity' : 'PEAK CAPACITY/HR', 
            'avgCapacity' : 'AVERAGE CAPACITY/HR', 
            'peakPerformance' : 'PEAK PERFORMANCE/HR', 
            'avgPerformance' : 'AVERAGE PERFORMANCE/HR',
            'machineType': 'MACHINE',
            'machinist': 'MACHINIST',
            'department': 'DEPARTMENT',
        })

        # Creating Factory and Floor Columns and also assigning None if empty values found
        df_merged["FACTORY"] = df_merged["LINE"].apply(
            lambda x: FACTORY_FLOOR_MAPPING.get(x, (None, None))[0] if pd.notna(x) and str(x).strip() != '' else None
        )

        df_merged["FLOOR"] = df_merged["LINE"].apply(
            lambda x: FACTORY_FLOOR_MAPPING.get(x, (None, None))[1] if pd.notna(x) and str(x).strip() != '' else None
        )

        # # Define the desired column order
        # column_order = [
        #     "EMPLOYEE ID", "EMPLOYEE NAME", "LINE", "FACTORY", "FLOOR", "SECTION", "DESIGNATION", 
        #     "CODE", "OPERATION", "TYPE", "SAM", 
        #     "PEAK CAPACITY/HR", "AVERAGE CAPACITY/HR", "PEAK PERFORMANCE/HR", "AVERAGE PERFORMANCE/HR", "MACHINE", 
        #     "STATUS", "AVERAGE CAPACITY", "PEAK CAPACITY", "AVERAGE PERFORMANCE", "PEAK PERFORMANCE",
        #     "MACHINIST", "DEPARTMENT",
        # ]
        # emp_fact_path = "EMP_FACT_LATEST.csv"
        # if os.path.exists(emp_fact_path):
        #     os.remove(emp_fact_path)
        #     print(f"Existing file {emp_fact_path} deleted.")
        # # Save to CSV with specific column order and no index
        # df_merged.to_csv(emp_fact_path, columns=column_order, index=False)

        return df_merged
    except Exception as e:
        logger.info(e)
        logger.info(f"Error in merging skill matrix and operations dataframes: {e}")



# Function to process the skill matrix data
def process_skill_matrix_data(df_skill_matrix):

    # Remove rows where type is blank (empty string or NaN)
    df_skill_matrix = df_skill_matrix[~df_skill_matrix['type'].isin(['', ' ', None]) & df_skill_matrix['type'].notna()]

    # Remove rows where 'operationCode' is blank (empty string or NaN)
    df_skill_matrix = df_skill_matrix.dropna(subset=["operationCode"])

    # df_skill_matrix = df_skill_matrix[df_skill_matrix["lines"].apply(lambda x: x != "[]")]

    # Convert 'lines' column from string to list (if not already a list)
    df_skill_matrix["lines"] = df_skill_matrix.loc[:, "lines"] = df_skill_matrix["lines"].apply(
        lambda x: ast.literal_eval(x) if isinstance(x, str) and x.startswith("[") else x
    )

    # Ensure all values in 'lines' are lists
    df_skill_matrix["lines"] = df_skill_matrix.loc[:, "lines"] = df_skill_matrix["lines"].apply(
        lambda x: x if isinstance(x, list) else [x]
    )

    # Apply the function
    df_skill_matrix["lines"] = df_skill_matrix["lines"].apply(keep_first_valid_line)

    # Drop rows where 'lines' is now empty (None)
    df_skill_matrix = df_skill_matrix.dropna(subset=["lines"]).reset_index(drop=True)

    # Remove rows where 'type' is blank (empty string or NaN)
    df_skill_matrix = df_skill_matrix[df_skill_matrix["type"].notna() & (df_skill_matrix["type"].str.strip() != "")]

    # Replace '9A' with '9' and '9B' with '10'
    df_skill_matrix["lines"] = df_skill_matrix["lines"].replace({"Line 9A": "Line 9", "Line 9B": "Line 10"})

    return df_skill_matrix



def export_to_excel(json_data, style):
    """
    Converts JSON data into an Excel file with two sheets:
    - "General Info" (includes Unique Styles, Buyers, and Machinist details)
    - "Table Data" (detailed table)

    Args:
        json_data (dict): Input JSON data
        req_body (dict, optional): Request body for fallback styles. Default is None.
        file_name (str, optional): Output Excel file name. Default is "output.xlsx".
    """

    # Use `unique_styles` if present, else fetch from `req_body["style"]`
    unique_styles = json_data.get("unique_styles")
    if not unique_styles and style:  # If unique_styles is missing, check req_body
        unique_styles = [style] if style else []  # Convert to list if it's a string

    # Column name mapping for Machinist & Non-Machinist Count
    column_mapping = {
        "machinist_available": "Machinist Available",
        "non_machinist_available": "Non Machinist Available",
        "machinist_required": "Machinist Required",
        "non_machinist_required": "Non Machinist Required",
        "total_required": "Total Required",
        "total_available": "Total Available"
    }

    # Create a new Excel workbook
    wb = Workbook()

    # ========== SHEET 1: "General Info" ==========
    ws_info = wb.active
    ws_info.title = "General Info"

    # Bold Font Style
    bold_font = Font(bold=True)

    row = 1  # Start at row 1

    # Writing Unique Styles (either from JSON or req.body)
    ws_info.cell(row=row, column=1, value="Unique Styles").font = bold_font
    row += 1
    if unique_styles:  # Only write if styles are available
        for style in unique_styles:
            ws_info.cell(row=row, column=1, value=style)
            row += 1
    row += 1  # Add a blank row for spacing

    # Writing Buyer Info (Value below the header)
    ws_info.cell(row=row, column=1, value="Buyers").font = bold_font
    row += 1
    ws_info.cell(row=row, column=1, value=json_data["info"]["buyers"])
    row += 2  # Move to next section

    # Writing Total Machinist Count Count
    ws_info.cell(row=row, column=1, value="Total Machinist Count").font = bold_font
    row += 1
    for key, value in json_data["machinist_nonMachinist_count"].items():
        display_name = column_mapping.get(key, key)  # Convert key if in mapping
        ws_info.cell(row=row, column=1, value=display_name)
        ws_info.cell(row=row, column=2, value=value)
        row += 1
    row += 1  # Blank row for spacing

    # Writing Machinist Info
    ws_info.cell(row=row, column=1, value="Machinist Info").font = bold_font
    row += 1
    for key, value in json_data["machinist_nonMachinist_info"].items():
        ws_info.cell(row=row, column=1, value=key)
        ws_info.cell(row=row, column=2, value=value)
        row += 1

    # Writing Prediction Report
    row += 1  # Add spacing before Prediction Report
    ws_info.cell(row=row, column=1, value="Prediction Report").font = bold_font
    row += 1
    prediction_data = json_data['Target data']
    for item in prediction_data:
        for section in item.get("production_target", []):
            ws_info.cell(row=row, column=1, value="Production Target")
            ws_info.cell(row=row, column=2, value=section["total_planned_qty"])
            row += 1
        for section in item.get("predicted_production", []):
            ws_info.cell(row=row, column=1, value="Predicted Production")
            ws_info.cell(row=row, column=2, value=section["total_planned_qty"])
            row += 1

    # Auto-adjust column width dynamically in "General Info"
    for col in range(1, 3):  # Adjusting first two columns
        max_length = 0
        for cell in ws_info[get_column_letter(col)]:  # Loop through each column
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws_info.column_dimensions[get_column_letter(col)].width = max_length + 5

    # ========== SHEET 2: "Table Data" ==========
    ws_table = wb.create_sheet(title="Table Data")

    # Convert table_data to DataFrame
    df = pd.DataFrame(json_data["table_data"])
    if 'Preferred Employees' in df.columns:
        df['Preferred Employees'] = df['Preferred Employees'].astype(str)

    # Write column headers with bold font
    row = 1
    for col_num, column_title in enumerate(df.columns, start=1):
        cell = ws_table.cell(row=row, column=col_num, value=column_title)
        cell.font = bold_font  # Make column headers bold

    row += 1  # Move to the next row after headers

    # Write table data row by row
    for index, record in df.iterrows():
        for col_num, value in enumerate(record, start=1):
            ws_table.cell(row=row, column=col_num, value=value)
        row += 1

    # Auto-adjust column width dynamically in "Table Data"
    for col_num, column_title in enumerate(df.columns, start=1):
        max_length = len(column_title)
        for row_num in range(2, row):  # Start from row 2 (skip header)
            cell_value = ws_table.cell(row=row_num, column=col_num).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        ws_table.column_dimensions[get_column_letter(col_num)].width = max_length + 5

    # Save Excel file in memory using BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)  # Move to the beginning of the BytesIO stream

    return output  # Returning the Excel file as a byte stream




def custom_round(value):
    """
    Custom rounding function:
    - If decimal part < 0.5 → round down (floor)
    - If decimal part >= 0.5 → round up (ceil)
    - Handles NaN and Infinity by returning 0
    """
    if isinstance(value, float):
        if math.isnan(value) or math.isinf(value):  # Handle NaN or Infinity
            return 0
        integer_part = math.floor(value)  # Get the base integer
        decimal_part = value - integer_part  # Extract the decimal part
        return math.floor(value) if decimal_part < 0.5 else math.ceil(value)
    return value  # Return as is if not a float



def export_json_to_excel(json_data):
    """
    Converts JSON data into an Excel file with each section as a separate sheet.
    Each sheet contains:
        - Unique Styles
        - Buyers
        - Machinist & Non-Machinist Count
        - Machinist & Non-Machinist Info
        - Table Data

    Args:
        json_data (dict): The JSON data to be saved.

    Returns:
        BytesIO: The Excel file stored in memory.
    """

    # Column name mapping for Machinist & Non-Machinist Count
    column_mapping = {
        "machinist_available": "Machinist Available",
        "non_machinist_available": "Non Machinist Available",
        "machinist_required": "Machinist Required",
        "non_machinist_required": "Non Machinist Required",
        "total_required": "Total Required",
        "total_available": "Total Available"
    }

    # Create a new workbook
    wb = Workbook()
    
    bold_font = Font(bold=True)

    # ========== CREATE SHEETS PER SECTION ==========
    for section, table_data in json_data.get("table_data", {}).items():
        ws = wb.create_sheet(title=section)

        row = 1  # Start at row 1

        # ========== Unique Styles ==========
        ws.cell(row=row, column=1, value="Unique Styles").font = bold_font
        row += 1
        styles = json_data.get("unique_styles", {}).get(section, {}).get("unique_styles", [])

        if isinstance(styles, list):
            for style in styles:
                ws.cell(row=row, column=1, value=style)
                row += 1
        else:
            ws.cell(row=row, column=1, value=str(styles))  # Handle non-list cases
        row += 1  

        # ========== Buyers ==========
        ws.cell(row=row, column=1, value="Buyers").font = bold_font
        row += 1
        buyers = json_data.get("info", {}).get(section, {}).get("buyers", [])

        if isinstance(buyers, list):
            for buyer in buyers:
                ws.cell(row=row, column=1, value=buyer)
                row += 1
        else:
            ws.cell(row=row, column=1, value=str(buyers))  # Handle non-list cases
        row += 1  

        # ========== Total Machinist Count Count ==========
        ws.cell(row=row, column=1, value="Total Machinist Count").font = bold_font
        row += 1
        count_data = json_data.get("machinist_nonMachinist_count", {}).get(section, {})
        for key, value in count_data.items():
            ws.cell(row=row, column=1, value=column_mapping.get(key, key))
            ws.cell(row=row, column=2, value=value)
            row += 1
        row += 1  

        # ========== Machinist Info ==========
        ws.cell(row=row, column=1, value="Machinist Info").font = bold_font
        row += 1
        info_data = json_data.get("machinist_nonMachinist_info", {}).get(section, {})
        for key, value in info_data.items():
            ws.cell(row=row, column=1, value=key)
            ws.cell(row=row, column=2, value=value)
            row += 1
        row += 1

        # ========== Prediction Report ==========
        ws.cell(row=row, column=1, value="Prediction Report").font = bold_font
        row += 1
        prediction_data = json_data.get("prediction_report", {}).get(section, {})
        report_data = prediction_data['Target data']
        for item in report_data:
            for section in item.get("production_target", []):
                ws.cell(row=row, column=1, value="Production Target")
                ws.cell(row=row, column=2, value=section["total_planned_qty"])
                row += 1
            for section in item.get("predicted_production", []):
                ws.cell(row=row, column=1, value="Predicted Production")
                ws.cell(row=row, column=2, value=section["total_planned_qty"])
                row += 1
        row += 2

        # ========== Table Data ==========
        if table_data:
            df = pd.DataFrame(table_data)
            ws.cell(row=row, column=1, value="Table Data").font = bold_font
            row += 1

            # Write column headers with bold font
            for col_num, column_title in enumerate(df.columns, start=1):
                cell = ws.cell(row=row, column=col_num, value=column_title)
                cell.font = bold_font  

            row += 1  # Move to the next row after headers

            # Write table data row by row
            for record in df.itertuples(index=False):
                for col_num, value in enumerate(record, start=1):
                    ws.cell(row=row, column=col_num, value=value)
                row += 1

            # Auto-adjust column width dynamically
            for col_num, column_title in enumerate(df.columns, start=1):
                max_length = max(len(str(column_title)), max(len(str(cell.value)) for cell in ws[get_column_letter(col_num)] if cell.value))
                ws.column_dimensions[get_column_letter(col_num)].width = max_length + 5

    # Remove the default empty sheet created by openpyxl
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # Save Excel file in memory
    output = BytesIO()
    wb.save(output)
    output.seek(0)  

    return output


# Push Notification Utilities

def create_bulk_push_notifications(notification_type, title, message, users, data=None):
    """
    Create push notifications for multiple users
    
    Args:
        notification_type: Type of notification
        title: Notification title  
        message: Notification message
        users: List of User objects
        data: Optional JSON data for the notification
        
    Returns:
        List of created notification objects
    """
    
    notifications = []
    with transaction.atomic():
        for user in users:
            notification = PushNotification(
                notification_type=notification_type,
                title=title,
                message=message,
                user=user,
                data=data
            )
            notifications.append(notification)
        
        # Bulk create all notifications at once
        return PushNotification.objects.bulk_create(notifications)

def get_notification_type_by_time():
    """
    Determine the appropriate notification type based on the current time.
    
    Returns:
        str: The notification type based on time ranges:
            - 'dday_8_50' for 8:30 AM to 12:00 PM
            - 'dday_12_45' for 12:45 PM to 5:00 PM
            - 'dday_5_30' for after 5:00 PM
    """
    current_time = datetime.now().time()
    
    # Define time ranges
    morning_start = datetime.strptime('08:50:00', '%H:%M:%S').time()
    morning_end = datetime.strptime('12:00:00', '%H:%M:%S').time()
    afternoon_start = datetime.strptime('12:45:00', '%H:%M:%S').time()
    afternoon_end = datetime.strptime('17:00:00', '%H:%M:%S').time()
    
    if morning_start <= current_time <= morning_end:
        return 'dday_8_50'
    elif afternoon_start <= current_time <= afternoon_end:
        return 'dday_12_45'
    else:
        return 'dday_5_30'


# Function to delete old exported files
def delete_old_exported_files(days_old, file_extension):
    """
    Deletes .xlsx files from the exports folder that are older than the specified number of days.

    Args:
        days_old (int): Number of days to consider a file old. Defaults to 7.
    """
    exports_dir = os.path.join(settings.BASE_DIR, 'exports')

    if not os.path.exists(exports_dir):
        logger.info("Exports directory does not exist.")
        return

    today = datetime.now().date()
    deleted_files = []

    for file in os.listdir(exports_dir):
        file_path = os.path.join(exports_dir, file)
        if file.endswith(file_extension) and os.path.isfile(file_path):
            modified_time = datetime.fromtimestamp(os.path.getmtime(file_path)).date()
            if (today - modified_time).days >= days_old:
                os.remove(file_path)
                deleted_files.append(file)

    logger.info(f"Deleted {len(deleted_files)} old file(s): {deleted_files}")

# Merge operations and style_ob dataframes
def merge_machine_sam(df_operations, df_sections):
    df_sections_unique = df_sections.drop_duplicates(subset='code')
    df_merged = df_operations.merge(
        df_sections_unique[['code', 'machineType', 'sam']],
        how='left',
        left_on='Operation Code',
        right_on='code'
    )
    df_merged.drop(columns=['code'], inplace=True)
    return df_merged


# Using Graph logic on dataframe
def build_graph(df):
    G = nx.DiGraph()
    op_name_to_code = dict(zip(df['Operation Name'], df['Operation Code']))

    for idx, row in df.iterrows():
        node_id = row['Operation Code']
        G.add_node(node_id)

    for idx, row in df.iterrows():
        node_id = row['Operation Code']
        dependencies = row['Dependencies']
        if pd.notna(dependencies):
            dependency_list = [d.strip() for d in dependencies.split(",")]
            for dep in dependency_list:
                dep_code = op_name_to_code.get(dep)
                if dep_code:
                    G.add_edge(dep_code, node_id)
    return G



def topological_sort_with_mitigation(G):
    try:
        return list(nx.topological_sort(G)), []
    except nx.NetworkXUnfeasible:
        cycles = list(nx.simple_cycles(G))
        broken_edges = []
        for cycle in cycles:
            if len(cycle) > 1:
                u, v = cycle[0], cycle[1]
                if G.has_edge(u, v):
                    G.remove_edge(u, v)
                    broken_edges.append((u, v))
        try:
            return list(nx.topological_sort(G)), broken_edges
        except:
            return [], broken_edges
        

# Processing styles
def process_styles(df_merged):
    final_records = []
    # mitigation_logs = []

    # Check if section exists, else create dummy section
    if 'section' not in df_merged.columns:
        if 'Section' in df_merged.columns:
            df_merged.rename(columns={'Section': 'section'}, inplace=True)
        else:
            df_merged['section'] = 'Default'

    for style_name, style_df in df_merged.groupby('Name'):
        style_df = style_df.copy()

        # Build graph for style
        G_style = build_graph(style_df)
        sorted_nodes, broken_edges_style = topological_sort_with_mitigation(G_style)
        overall_sequence_map = {code: seq+1 for seq, code in enumerate(sorted_nodes)}
        style_df['Operation Sequence (Overall)'] = style_df['Operation Code'].map(overall_sequence_map)

        # if broken_edges_style:
        #     mitigation_logs.append(f"Mitigated cycles in style: {style_name} (global)")

        # Section-wise sequencing
        section_sequence_list = []
        for section_name, group in style_df.groupby('section'):
            G_section = build_graph(group)
            sorted_section_nodes, broken_edges_section = topological_sort_with_mitigation(G_section)
            section_seq_map = {code: seq+1 for seq, code in enumerate(sorted_section_nodes)}

            temp_df = group.copy()
            temp_df['Operation Sequence (Section wise)'] = temp_df['Operation Code'].map(section_seq_map)
            section_sequence_list.append(temp_df)

            # if broken_edges_section:
            #     mitigation_logs.append(f"Mitigated cycles in style: {style_name} - section: {section_name}")

        style_final_df = pd.concat(section_sequence_list)
        final_records.append(style_final_df)

    df_final = pd.concat(final_records)
    return df_final


# Renaming columns for style_ob
def renaming_columns_style_ob(df_final):

    # Rename columns before saving
    rename_mapping = {
        'Name': 'style',
        'Department': 'department',
        'Operation Name': 'operation',
        'Operation Code': 'code',
        'Operation Type': 'type',
        'machineType': 'machine_type',
        'Machinist': 'machinist',
        'Operation Sequence (Section wise)': 'op_seq',
        'Created At': 'created_at'
    }

    df_final.rename(columns=rename_mapping, inplace=True)

    df_final.sort_values(by=['style', 'section', 'op_seq'], inplace=True)
    df_final['style'] = df_final['style'].apply(clean_string)

    return df_final


# Cleaning function
def clean_string(s):
    if not isinstance(s, str):
        return ''
    s = s.strip().lower()
    s = re.sub(r'\s+', '', s)

    # Check if there are unbalanced parentheses
    if s.count('(') > s.count(')'):
        s += ')'
    return s


# Remove employee from preferred_employees based on final_allocation
def clean_preferred_employees(df):
    def remove_employee(row):
        final_allocation = row['final_allocation']
        preferred = row['preferred_employees']
        if pd.notna(final_allocation) and isinstance(preferred, str):
            # Extract employee name from final_allocation
            match = re.search(r'-\s+(.*?)\s+\(Line', final_allocation)
            if match:
                allocated_name = match.group(1).strip()
                
                # Create a pattern to match the specific employee entry
                # Pattern: employee_name - employee_id [Line: ...]
                pattern = rf'\b{re.escape(allocated_name)}\s+-\s+\d+\s+\[Line:[^\]]+\]'
                
                # Remove the matched employee entry
                updated_preferred = re.sub(pattern, '', preferred)
                
                # Clean up any resulting issues with commas and separators
                # Remove double commas, leading/trailing commas in sections
                updated_preferred = re.sub(r',\s*,', ',', updated_preferred)  # Double commas
                updated_preferred = re.sub(r':\s*,', ':', updated_preferred)   # Comma right after colon
                updated_preferred = re.sub(r',\s*\|', ' |', updated_preferred) # Comma before |
                updated_preferred = re.sub(r'\|\s*,', '|', updated_preferred)  # Comma after |
                
                # Remove empty categories (e.g., "SAME FACTORY: |" or "SAME FACTORY:")
                # This pattern matches category labels followed by colon and optional whitespace, 
                # then either end of string or pipe separator
                updated_preferred = re.sub(r'[A-Z\s]+:\s*(?=\||$)', '', updated_preferred)
                
                # Clean up multiple pipes and leading/trailing pipes
                updated_preferred = re.sub(r'\|\s*\|', '|', updated_preferred)  # Multiple pipes
                updated_preferred = re.sub(r'^\s*\|\s*', '', updated_preferred)  # Leading pipe
                updated_preferred = re.sub(r'\s*\|\s*$', '', updated_preferred)  # Trailing pipe
                
                # Clean up extra whitespace
                updated_preferred = re.sub(r'\s+', ' ', updated_preferred).strip()
                
                return updated_preferred
        return preferred
    
    df['preferred_employees'] = df.apply(remove_employee, axis=1)
    return df


def fetch_dday_data(line_no):
    # Define column mapping for consistent naming
    column_mapping = {
        'id': 'Dday_ID',
        'factory': 'Factory',
        'order_no': 'Token No',
        'style': 'Style',
        'allocated_emp_name': 'Allocated Employee Name',
        'allocated_emp_id': 'Allocated Employee ID',
        'original_emp_name': 'Operator Name',
        'original_emp': 'Operator ID',
        'line': 'Original Line',
        'reallocation_reason': 'Reallocation Reason',
        'shortage_reason': 'Shortage Reason',
        'designation': 'M/c Or Nm/c',
        'code': 'Operational code',
        'operation': 'Core Operation',
        'sam': 'SAM',
        'smv': 'SMV',
        'section': 'Section',
        'target_100': 'Planned Allocated Quantity',
        'allocated_capacity': 'Allocated Capacity',
        'allocated_frm_line': 'Allocated line',
        'attendance_status': 'Attendance Status',
        'final_allocation': 'Final Allocation',
        'preferred_employees': 'Preferred Employees',
        'wip_quantity': 'WIP Qty',
    }

    # Only select necessary columns to improve query performance
    required_fields = list(column_mapping.keys())

    # Optimize database query using values() with specific fields and defer sorting to the Python side
    if line_no == 'All':
        queryset = DDayData.objects.values(*required_fields)
    else:
        queryset = DDayData.objects.filter(line=line_no).values(*required_fields)

    # Execute query and get results in a list
    filtered_data = list(queryset)

    # Return empty template if no data found
    if not filtered_data:
        empty_data = {key: "" for key in column_mapping.values()}
        return {"message": 'No data to display', "data": {"table_data": empty_data},  "status": status.HTTP_200_OK}

    
    # Process data efficiently using pandas
    df = pd.DataFrame(filtered_data)
    
    # Handle timezone-aware datetimes more efficiently
    datetime_cols = df.select_dtypes(include=['datetime64[ns, UTC]', 'datetimetz']).columns
    for col in datetime_cols:
        df[col] = df[col].dt.tz_localize(None)
    
    # Replace NaN values with None for JSON serialization
    df = df.replace({np.nan: None})
    # Handle infinity values separately to avoid potential errors
    for col in df.columns:
        if df[col].dtype == 'float64' or df[col].dtype == 'float32':
            df[col] = df[col].replace([np.inf, -np.inf], None)
    
    # Sort by allocated capacity in descending order
    df = df.sort_values(by='allocated_capacity', ascending=False)

    # Remove employees from preferred_employees based on final_allocation
    df = clean_preferred_employees(df)
    
    df = df.rename(columns=column_mapping)

    return {"message": "Success", "data": {"records": df.to_dict(orient="records")},  "status": status.HTTP_200_OK}



def fetch_attendance_data(line_no, today, yesterday):
    # Create datetime ranges for today and yesterday
    today_start = datetime.combine(today, datetime.min.time())
    today_end = datetime.combine(today, datetime.max.time())
    yesterday_start = datetime.combine(yesterday, datetime.min.time())
    yesterday_end = datetime.combine(yesterday, datetime.max.time())
    
    # Build filter conditions - fixing the date lookup issue
    filters = Q(attendance_date__range=(yesterday_start, today_end))
    if line_no != 'All':
        filters &= Q(line=line_no)

    # This single query calculates all the attendance metrics we need
    attendance_metrics = AttendanceMaster.objects.filter(filters).aggregate(
        # Today's attendance
        present_today=Sum(
            Case(
                When(attendance_date__gte=today_start, attendance_date__lte=today_end, status='P', then=Value(1)),
                default=Value(0),
                output_field=IntegerField()
            )
        ),
        absent_today=Sum(
            Case(
                When(attendance_date__gte=today_start, attendance_date__lte=today_end, status='A', then=Value(1)),
                default=Value(0),
                output_field=IntegerField()
            )
        ),
        # Yesterday's attendance
        present_yesterday=Sum(
            Case(
                When(attendance_date__gte=yesterday_start, attendance_date__lte=yesterday_end, status='P', then=Value(1)),
                default=Value(0),
                output_field=IntegerField()
            )
        ),
        absent_yesterday=Sum(
            Case(
                When(attendance_date__gte=yesterday_start, attendance_date__lte=yesterday_end, status='A', then=Value(1)),
                default=Value(0),
                output_field=IntegerField()
            )
        )
    )
    
    # Handle None values that can occur if no records match
    present = attendance_metrics['present_today'] or 0
    absent = attendance_metrics['absent_today'] or 0
    present_yesterday = attendance_metrics['present_yesterday'] or 0
    absent_yesterday = attendance_metrics['absent_yesterday'] or 0
    
    # Calculate differences
    present_diff = present - present_yesterday
    absent_diff = absent - absent_yesterday
    
    # Prepare response data using dict comprehension for cleaner code
    attendance_stats = {
        status: {
            "count": abs(diff),
            "direction": "increase" if diff > 0 else ("decrease" if diff < 0 else "no change")
        }
        for status, diff in [("present", present_diff), ("absent", absent_diff)]
    }
    
    attendance = {
        "Planned Attendance": present + absent,
        "Present": present,
        "Absent": absent,
    }
    return {"message": "Success", "data": {"attendance_data": attendance, "attendance_stats": attendance_stats},  "status": status.HTTP_200_OK}


# Remove employee by their ID from a list of employee dictionaries
def remove_by_employee_id(employee_list, emp_id_to_remove):
    """Remove employee by their ID"""
    return [emp_dict for emp_dict in employee_list if emp_id_to_remove not in emp_dict]


# Function to transform unallocated employees to on hold and add in EmployeesOnHold model
def transform_unallocated_to_on_hold_from_dict(unallocated_data_dicts):
    grouped_data = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))

    for ue in unallocated_data_dicts:
        key_line = ue.get('line') or 'UnknownLine'
        key_section = ue.get('section') or 'UnknownSection'
        key_date = ue.get('date')
        employee_id = str(ue.get('employee_id')) if ue.get('employee_id') is not None else "UnknownID"
        employee_name = ue.get('employee_name') or "Unknown"
        employee_dict = {employee_id: employee_name}
        grouped_data[key_line][key_section][key_date].append(employee_dict)

    # Collect all new EmployeesOnHold objects
    new_entries = []

    for line, sections in grouped_data.items():
        for section, dates in sections.items():
            for date, employees_list in dates.items():
                preferred_employees_json = json.dumps(employees_list)
                count = len(employees_list)
                new_entries.append(EmployeesOnHold(
                    line=line,
                    section=section,
                    date=date,
                    preferred_employees=preferred_employees_json,
                    count=count
                ))

    # Bulk insert with optional truncation
    with transaction.atomic():
        truncate_table(EmployeesOnHold)
        EmployeesOnHold.objects.bulk_create(new_entries, batch_size=1000)

    logger.info(f"Inserted {len(new_entries)} EmployeesOnHold records.")


# To add code also but on standby as of now
def transform_unallocated_to_on_hold_from_dict_updated(unallocated_data_dicts):
    grouped_data = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: defaultdict(list))))
    for ue in unallocated_data_dicts:
        key_line = ue.get('line') or 'UnknownLine'
        key_section = ue.get('section') or 'UnknownSection'
        key_code = ue.get('code') or 'UnknownCode'
        key_date = ue.get('date')
        employee_id = str(ue.get('employee_id')) if ue.get('employee_id') is not None else "UnknownID"
        employee_name = ue.get('employee_name') or "Unknown"
        employee_dict = {employee_id: employee_name}
        grouped_data[key_line][key_section][key_code][key_date].append(employee_dict)
    
    # Collect all new EmployeesOnHold objects
    new_entries = []
    for line, sections in grouped_data.items():
        for section, codes in sections.items():
            for code, dates in codes.items():
                for date, employees_list in dates.items():
                    preferred_employees_json = json.dumps(employees_list)
                    count = len(employees_list)
                    new_entries.append(EmployeesOnHold(
                        line=line,
                        section=section,
                        code=code,
                        date=date,
                        preferred_employees=preferred_employees_json,
                        count=count
                    ))
    
    # Bulk insert with optional truncation
    with transaction.atomic():
        truncate_table(EmployeesOnHold)
        EmployeesOnHold.objects.bulk_create(new_entries, batch_size=1000)
    logger.info(f"Inserted {len(new_entries)} EmployeesOnHold records.")



def update_sections(input_array, main_array):
    if not input_array:
        return [{'section': section, 'total_planned_qty': 0} for section in main_array]

    existing_sections = {item['section'] for item in input_array}
    missing_sections = [section for section in main_array if section not in existing_sections]

    input_array.extend({'section': section, 'total_planned_qty': 0} for section in missing_sections)
    return input_array



def remove_duplicate_employee_dicts(data):
    seen = {}
    for entry in data:
        for emp_id, name in entry.items():
            seen[emp_id] = name  # will overwrite duplicates, keeping the last occurrence
    return [{emp_id: name} for emp_id, name in seen.items()]




def fetchMaxQtyDday(df):

    latest_ops = (
        df.groupby(['style', 'line', 'section', 'code'])['op_seq']
        .max()
        .reset_index()
        .rename(columns={'op_seq': 'max_op_seq'})
    )

    df_latest_ops = pd.merge(df, latest_ops, on=['style', 'line', 'section', 'code'], how='inner')
    df_latest_ops = df_latest_ops[
        (df_latest_ops['op_seq'] == df_latest_ops['max_op_seq']) &
        (df_latest_ops['section'].str.lower().str.contains('assembly'))
    ]

    grouped_qty_stats = (
        df_latest_ops.groupby(['planned_dates', 'line', 'style', 'section', 'code'])
        .agg({
            'planned_qty': ['min', 'max', 'sum'],
            'allocated_capacity': ['min', 'max', 'sum'],
            'op_seq': 'first'
        })
        .reset_index()
    )

    grouped_qty_stats.columns = [
        'planned_dates', 'line', 'style', 'section', 'code',
        'min_planned_qty', 'max_planned_qty', 'sum_planned_qty',
        'min_allocated_capacity', 'max_allocated_capacity', 'sum_allocated_capacity',
        'op_seq'
    ]

    top_capacity_per_line_style = (

        grouped_qty_stats.sort_values('sum_allocated_capacity', ascending=False)

        .groupby(['line','style','section'], as_index=False)

        .first()

    )

    return top_capacity_per_line_style