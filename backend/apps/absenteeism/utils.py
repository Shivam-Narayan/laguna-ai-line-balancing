import io
import math
import json
import base64
import logging
import pandas as pd

from io import BytesIO
from openpyxl import Workbook
from datetime import datetime
from django.conf import settings
from collections import defaultdict
from django.http import HttpResponse
from sendgrid import SendGridAPIClient
from openpyxl.utils import get_column_letter
from django.template.loader import render_to_string
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from sendgrid.helpers.mail import Mail, Attachment, FileContent, FileName, FileType, Disposition

from apps.absenteeism.models import AbsenteeismPrediction
from apps.data_engine.models import LocalHolidayCalendar, PayableWorkingDays

logger = logging.getLogger('general')


def generate_csv():
    """
    Generate CSV data from the LocalHolidayCalendar model and return it in memory.
    """
    try:
        fields = [field.name for field in LocalHolidayCalendar._meta.get_fields()]
        queryset = LocalHolidayCalendar.objects.all()
        data = list(queryset.values(*fields))

        if not data:
            return None

        df = pd.DataFrame(data)
        csv_data = io.StringIO()
        df.to_csv(csv_data, index=False)
        csv_data.seek(0)  # Rewind the StringIO object to the beginning for reading

        return csv_data

    except Exception as e:
        return None


# SendGrid API configuration
def send_email(recipient_emails, data, subject, type, file_name, test=False):
    try:
        logger.info(f"*******************************************************************")
        logger.info(f"Running EMAIL FUNCTION at {str(datetime.now())} hours!")
        # Get today's date
        # Render the email body from the template
        email_body = render_to_string('email_template.html', {'subject': subject})

        # SendGrid API configuration
        sendgrid_api_key = settings.SENDGRID_API_KEY
        sendgrid_from_email = settings.DEFAULT_FROM_EMAIL

        # Ensure recipient_emails is a list
        if isinstance(recipient_emails, str):
            recipient_emails = recipient_emails.split(',')  # Split comma-separated string into a list

        # Create the Mail object
        message = Mail(
            from_email=sendgrid_from_email,
            to_emails=recipient_emails,  # Pass the list of email addresses
            subject=subject,
            html_content=email_body,
        )

        attachments_list = []

        # Add attachment if data is provided
        if data:
            # Base64 encode the file (SendGrid requires this format)
            encoded_file = base64.b64encode(data.getvalue()).decode()

            # Create the attachment
            attached_file = Attachment(
                file_content=FileContent(encoded_file),
                file_name=FileName(file_name),
                file_type=FileType(type),
                disposition=Disposition('attachment')
            )
            # message.attachment = attached_file
            attachments_list.append(attached_file)

        if test:
            with open('csv_files/attendance.csv', 'rb') as f:
                encoded_csv = base64.b64encode(f.read()).decode()
                attached_file_2 = Attachment(
                    file_content=FileContent(encoded_csv),
                    file_name=FileName("attendance.csv"),
                    file_type=FileType("text/csv"),
                    disposition=Disposition('attachment')
                )
                attachments_list.append(attached_file_2)

        message.attachment = attachments_list

        # Send the email
        sg = SendGridAPIClient(sendgrid_api_key)
        response = sg.send(message)
        logger.info(f"Response status code {response.status_code}!")

        # Check for successful response
        if response.status_code in [200, 202]:
            return email_body
        else:
            logger.info(f"Error sending email: {response.status_code}, {response.body}")
            return None

    except Exception as e:
        logger.info(f"Error sending email: {e}")
        return None

def generate_prediction_data(data):
    sections = [
        {
            "Section": op["section"].upper(),
            "Total Active Operators": op["count"],
            "Total Operators Present": supply["count"],
            "Total Operatos Gap": gap["count"],
        }
        for op, supply, gap in zip(data["total_operators"], data["total_operators_supply"], data["total_operators_gap"])
    ]

    # Add the "Total" row
    sections.append({
        "Section": "Total",
        "Total Active Operators": data['total_employees'],
        "Total Operators Present": data['projected_attendance'],
        "Total Operatos Gap": data['total_predicted_absenteeism'],
    })
    
    # Add the "Absenteeism percentage" row
    sections.append({
        "Section": "Predicted Absenteeism (%)",
        "Total Active Operators": data['absenteeism_percentage'],
        "Total Operators Present": "",
        "Total Operatos Gap": "",
    })

    line = f"{data['line']} lines" if data['line'].lower() == 'all' else data['line']
    forecast_period = data['forecast_period']

    excel_data = convert_to_excel_data(sections, "Absenteeism Prediction Report", data["Target data"], prediction=[line, forecast_period], machinists_info=[data["required_machinists"], data["actual_machinists"]])

    return excel_data    
    
    
def convert_number(value):
    if value - math.floor(value) < 0.5:
        return math.floor(value)
    else:
        return math.ceil(value)



def convert_to_excel_data(data, sheet_name, shortage_data=None, prediction=None, machinists_info=None):
    """
    Convert a list of dictionaries into an Excel file with auto-adjusted column widths.
    
    If prediction is provided, adds a heading and writes main data starting at row 4.
    If shortage_data is provided, appends "Predicted Production Shortage" section below main table.
    
    Args:
        data (list): List of dictionaries containing the main data.
        shortage_data (list): List containing production and prediction data.
        prediction (tuple): (some_label, some_date) tuple for heading.
        
    Returns:
        BytesIO: The Excel file as a binary stream.
    """
    excel_data = BytesIO()
    df = pd.DataFrame(data)

    with pd.ExcelWriter(excel_data, engine="openpyxl") as writer:
        worksheet = None

        if prediction is not None:
            # Write heading in row 1, merged across columns (number of columns = df columns count)
            heading_col_span = len(df.columns)
            heading_text = f"Absenteeism Prediction Report for {prediction[0]} and forecast period {prediction[1]}"
            
            # Create empty sheet first (will add data later)
            # Write blank sheet to get worksheet object
            df_empty = pd.DataFrame()
            df_empty.to_excel(writer, index=False, sheet_name=sheet_name)
            worksheet = writer.sheets[sheet_name]

            worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=heading_col_span)
            heading_cell = worksheet.cell(row=1, column=1, value=heading_text)
            heading_cell.font = Font(bold=True)
            heading_cell.alignment = Alignment(horizontal="center")

            # Now write the main data starting at row 4 (skip rows 2 and 3)
            for r_idx, row in enumerate(df.itertuples(index=False), start=4):
                for c_idx, value in enumerate(row, start=1):
                    worksheet.cell(row=r_idx, column=c_idx, value=value)

            # Write column headers at row 3 (right above data)
            for c_idx, col_name in enumerate(df.columns, start=1):
                cell = worksheet.cell(row=3, column=c_idx, value=col_name)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")

        else:
            # If no prediction heading, just write data normally at row 1
            df.to_excel(writer, index=False, sheet_name=sheet_name)
            worksheet = writer.sheets[sheet_name]

        # Add shortage_data section if present
        if shortage_data is not None:
            # Calculate where main data ends
            if prediction is not None:
                start_row = 4 + len(df) + 3  # 3 rows below main data starting at row 4
            else:
                start_row = len(df) + 5  # default if no prediction heading

            heading_text = "Predicted Production Shortage"
            heading_col_span = 3

            worksheet.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=heading_col_span)
            heading_cell = worksheet.cell(row=start_row, column=1, value=heading_text)
            heading_cell.font = Font(bold=True)
            heading_cell.alignment = Alignment(horizontal="center")

            headers = ["Section", "Predicted Production", "Production Target"]
            for col_idx, header in enumerate(headers, start=1):
                cell = worksheet.cell(row=start_row + 2, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")

            entry = shortage_data[0]
            predicted_dict = {item["section"]: item["total_planned_qty"] for item in entry["predicted_production"]}
            target_dict = {item["section"]: item["total_planned_qty"] for item in entry["production_target"]}
            sections = sorted(set(predicted_dict.keys()).union(target_dict.keys()))

            for i, section in enumerate(sections):
                worksheet.cell(row=start_row + 3 + i, column=1, value=section.upper())
                worksheet.cell(row=start_row + 3 + i, column=2, value=predicted_dict.get(section, 0))
                worksheet.cell(row=start_row + 3 + i, column=3, value=target_dict.get(section, 0))

            # Auto-adjust columns for this shortage_data section
            for col_idx in range(1, heading_col_span + 1):
                values = [worksheet.cell(row=start_row + 2, column=col_idx).value]
                for i in range(len(sections)):
                    values.append(worksheet.cell(row=start_row + 3 + i, column=col_idx).value)
                max_width = max(len(str(val)) for val in values if val is not None)
                worksheet.column_dimensions[get_column_letter(col_idx)].width = max_width + 2


        if machinists_info is not None:
            required_machinists_list = machinists_info[0]
            actual_machinists_list = machinists_info[1]
    
            last_row = worksheet.max_row

            start_row = last_row + 3  # Start new section after the last used row

            heading_text = "Machinist Information"
            heading_col_span = 3

            worksheet.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=heading_col_span)
            heading_cell = worksheet.cell(row=start_row, column=1, value=heading_text)
            heading_cell.font = Font(bold=True)
            heading_cell.alignment = Alignment(horizontal="center")

            headers = ["Section", "Required Machinists", "Actual Machinists"]
            for col_idx, header in enumerate(headers, start=1):
                cell = worksheet.cell(row=start_row + 2, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")

            required_machinists_dict = {item["section"]: item["count"] for item in required_machinists_list}
            actual_machinists_dict = {item["section"]: item["count"] for item in actual_machinists_list}
            sections = sorted(set(required_machinists_dict.keys()).union(actual_machinists_dict.keys()))

            for i, section in enumerate(sections):
                worksheet.cell(row=start_row + 3 + i, column=1, value=section.upper())
                worksheet.cell(row=start_row + 3 + i, column=2, value=required_machinists_dict.get(section, 0))
                worksheet.cell(row=start_row + 3 + i, column=3, value=actual_machinists_dict.get(section, 0))

            # Auto-adjust columns for this shortage_data section
            for col_idx in range(1, heading_col_span + 1):
                values = [worksheet.cell(row=start_row + 2, column=col_idx).value]
                for i in range(len(sections)):
                    values.append(worksheet.cell(row=start_row + 3 + i, column=col_idx).value)
                max_width = max(len(str(val)) for val in values if val is not None)
                worksheet.column_dimensions[get_column_letter(col_idx)].width = max_width + 2

        # Auto-adjust columns width for main data table (if exists)
        if df.shape[0] > 0:
            for col_idx, col_name in enumerate(df.columns, start=1):
                # Collect all values from data (including column header)
                if prediction is not None:
                    # Data starts at row 4, headers at 3
                    values = [col_name]
                    for r in range(4, 4 + len(df)):
                        val = worksheet.cell(row=r, column=col_idx).value
                        values.append(val)
                else:
                    values = [col_name] + list(df[col_name].astype(str))

                max_length = max(len(str(v)) for v in values if v is not None)
                worksheet.column_dimensions[get_column_letter(col_idx)].width = max_length + 2

    excel_data.seek(0)
    return excel_data


def calculate_absenteeism_percentage(df_active_emp, df_absentism):
    """
    Calculates absenteeism percentage per line per section grouped by year and month,
    including outlier removal.
 
    Args:
        df_active_emp (pd.DataFrame): DataFrame containing active employee data.
        df_absentism (pd.DataFrame): DataFrame containing absenteeism data.
 
    Returns:
        pd.DataFrame: DataFrame with absenteeism percentage.
    """
 
    emp_count_per_line_section = df_active_emp.groupby(['Line', 'Section'])['Emp No'].nunique().reset_index()
    emp_count_per_line_section.rename(columns={'Emp No': 'Total_Employees'}, inplace=True)

    
    absentism_count_per_month = df_absentism[df_absentism['Status'] == 'A'].groupby(['Year', 'Month', 'Line', 'Section'])['Status'].count().reset_index()
    absentism_count_per_month.rename(columns={'Status': 'Absent_Count'}, inplace=True)
 
 
    Q1 = absentism_count_per_month['Absent_Count'].quantile(0.25)
    Q3 = absentism_count_per_month['Absent_Count'].quantile(0.75)
    IQR = Q3 - Q1
    lower_bound = Q1 - 1.5 * IQR
    upper_bound = Q3 + 1.5 * IQR
    absentism_count_no_outliers = absentism_count_per_month[
        (absentism_count_per_month['Absent_Count'] >= lower_bound) &
        (absentism_count_per_month['Absent_Count'] <= upper_bound)
    ]
 
    final_data = pd.merge(absentism_count_no_outliers, emp_count_per_line_section, on=['Line', 'Section'], how='left')
    final_data['Absenteeism_Percentage'] = (final_data['Absent_Count'] / (final_data['Total_Employees']*26)) * 100
    final_data['Absenteeism_Percentage'] = round(final_data['Absenteeism_Percentage'],1)

    return final_data


def update_sections(input_array, main_array):
    main_sections = {item['section'] for item in main_array}  # Extract existing sections
    
    for item in input_array:
        if item['section'] not in main_sections:
            main_array.append({'section': item['section'], 'count': 0})
    
    return main_array


def merge_duplicates(lst, key="section", value="count"):
    merged_data = defaultdict(int)
    
    for item in lst:
        normalized_key = item[key].strip()  # Normalize by stripping spaces
        merged_data[normalized_key] += item[value]

    return [{key: k, value: v} for k, v in merged_data.items()]



# Function to check if the given date is a working day
def is_allowed_working_day(date_obj):

    # Get all payable working days from the database for the date range
    payable_dates_qs = list(PayableWorkingDays.objects.all().values())
    payable_dates = set()
    if payable_dates_qs:
        payable_dates_df = pd.DataFrame(payable_dates_qs)
        if 'date' in payable_dates_df.columns:
            payable_dates_df['date'] = pd.to_datetime(payable_dates_df['date']).dt.date
            payable_dates = set(payable_dates_df['date'])

    if date_obj in payable_dates:
        return True, "Allowed working day"

    # Load the LocalHolidayCalendar data from the database
    holiday_qs = list(LocalHolidayCalendar.objects.all().values())
    holiday_dates = set()
    if holiday_qs:
        local_holiday_calender_df = pd.DataFrame(holiday_qs)
        if 'date' in local_holiday_calender_df.columns:
            local_holiday_calender_df['date'] = pd.to_datetime(local_holiday_calender_df['date']).dt.date
            holiday_dates = set(local_holiday_calender_df['date'])

    if date_obj in holiday_dates:
        return False, "Local Holiday"

    # Sunday logic
    if date_obj.weekday() == 6:  # 6 represents Sunday
        return False, "Sunday"

    # Saturday logic
    if date_obj.weekday() == 5:  # 5 represents Saturday
        saturdays = [
            d.date() for d in pd.date_range(date_obj.replace(day=1), date_obj.replace(day=28) + pd.DateOffset(days=4))
            if d.weekday() == 5 and d.month == date_obj.month
        ]

        # Check if it's the 2nd, 3rd, or 4th Saturday
        if date_obj in saturdays:
            index = saturdays.index(date_obj)
            if index in [1, 2, 3]:  # 2nd, 3rd, or 4th Saturday
                suffix = "th"
                if index == 1:
                    suffix = "nd"
                elif index == 2:
                    suffix = "rd"
                return False, f"{index + 1}{suffix} Saturday"

    return True, "Allowed working day"

# Sum the counts of sections in a list of dictionaries
def sum_section_counts(data):
    section_totals = {}

    # Aggregate counts per section
    for entry_list in data:
        for item in entry_list:
            section = item['section']
            count = item['count']
            section_totals[section] = section_totals.get(section, 0) + count

    # Convert the result to list of dicts
    result = [{'section': section, 'count': count} for section, count in section_totals.items()]
    return result

# Normalize sections based on a predefined order, filling missing sections with count 0
def normalize_sections(data, section_order):
    section_map = {item['section']: item['count'] for item in data}

    normalized = []
    for section in section_order:
        count = section_map.get(section, 0)
        normalized.append({'section': section, 'count': count})

    return normalized



def write_absenteeism_data_to_excel(data, filename):
    """
    Write manufacturing absenteeism data to Excel with multiple sheets and formatting.
    
    Args:
        data: Dictionary containing absenteeism data for different production lines
        filename: Output Excel filename
    """
    
    # If data is a string (JSON), parse it
    if isinstance(data, str):
        data = json.loads(data)
    
    # Create a new workbook
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    def set_column_widths(worksheet, min_widths=None):
        """Helper function to set appropriate column widths"""
        if min_widths is None:
            min_widths = {}
            
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
                except:
                    pass
            
            # Apply minimum width if specified for this column
            header_cell = column[0]
            column_name = str(header_cell.value) if header_cell.value else ""
            
            # Set minimum widths for specific columns
            if column_name in min_widths:
                adjusted_width = max(min_widths[column_name], max_length + 2)
            else:
                adjusted_width = max_length + 2
                
            # Cap maximum width
            adjusted_width = min(adjusted_width, 25)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # 1. Summary Sheet
    ws_summary = wb.create_sheet("Summary")
    
    summary_data = []
    for line, line_data in data.items():
        total_employees = sum(line_data['total_employee_count'].values())
        total_predicted = sum([item['count'] for item in line_data['predicted_absent_count']])
        total_actual = sum([item['count'] for item in line_data['actual_absent_count']])
        predicted_avg = line_data.get('predicted_absenteeism_percentage', 0)
        actual_avg = line_data.get('actual_absenteeism_percentage', 0)
        
        # # Calculate weighted averages for percentages
        # predicted_avg = sum([
        #     line_data['predicted_absenteeism_percentage'].get(section, 0) * count 
        #     for section, count in line_data['total_employee_count'].items()
        # ]) / total_employees if total_employees > 0 else 0
        
        # actual_avg = sum([
        #     line_data['actual_absenteeism_percentage'].get(section, 0) * count 
        #     for section, count in line_data['total_employee_count'].items()
        # ]) / total_employees if total_employees > 0 else 0
        
        summary_data.append({
            'Lines': line,
            'Total Employees': total_employees,
            'Predicted Absent': int(total_predicted),
            'Actual Absent': total_actual,
            'Predicted Absenteeism %': round(predicted_avg, 2),
            'Actual Absenteeism %': round(actual_avg, 2),
            # 'Variance (Actual - Predicted)': total_actual - int(total_predicted),
            # 'Accuracy %': round(100 - abs(actual_avg - predicted_avg), 2)
        })
    
    df_summary = pd.DataFrame(summary_data)
    
    # Write summary to sheet
    for r in dataframe_to_rows(df_summary, index=False, header=True):
        ws_summary.append(r)
    
    # Format summary sheet
    for cell in ws_summary[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    # Set column widths with specific minimums for summary sheet
    summary_min_widths = {
        'Predicted Absent': 16,
        'Predicted Absenteeism %': 22,
        'Actual Absenteeism %': 20,
        'Variance (Actual - Predicted)': 25,
        'Accuracy %': 12
    }
    set_column_widths(ws_summary, summary_min_widths)
    
    # # 2. Detailed Data Sheet
    # ws_detailed = wb.create_sheet("Detailed Data")
    
    # detailed_data = []
    # for line, line_data in data.items():
    #     for section in line_data['total_employee_count'].keys():
    #         # Get predicted and actual counts for this section
    #         predicted_count = next((item['count'] for item in line_data['predicted_absent_count'] 
    #                               if item['section'] == section), 0)
    #         actual_count = next((item['count'] for item in line_data['actual_absent_count'] 
    #                            if item['section'] == section), 0)
            
    #         detailed_data.append({
    #             'Line': line,
    #             'Section': section,
    #             'Total Employees': line_data['total_employee_count'][section],
    #             'Predicted Absent Count': int(predicted_count),
    #             'Actual Absent Count': actual_count,
    #             'Predicted Absenteeism %': line_data['predicted_absenteeism_percentage'].get(section, 0),
    #             'Actual Absenteeism %': line_data['actual_absenteeism_percentage'].get(section, 0),
    #             'Count Variance': actual_count - int(predicted_count),
    #             'Percentage Variance': round(
    #                 line_data['actual_absenteeism_percentage'].get(section, 0) - 
    #                 line_data['predicted_absenteeism_percentage'].get(section, 0), 2
    #             )
    #         })
    
    # df_detailed = pd.DataFrame(detailed_data)
    
    # # Write detailed data to sheet
    # for r in dataframe_to_rows(df_detailed, index=False, header=True):
    #     ws_detailed.append(r)
    
    # # Format detailed sheet
    # for cell in ws_detailed[1]:
    #     cell.font = header_font
    #     cell.fill = header_fill
    #     cell.alignment = Alignment(horizontal='center')
    #     cell.border = border
    
    # # Set column widths with specific minimums for detailed sheet
    # detailed_min_widths = {
    #     'Predicted Absent Count': 20,
    #     'Actual Absent Count': 18,
    #     'Predicted Absenteeism %': 22,
    #     'Actual Absenteeism %': 20,
    #     'Total Employees': 15
    # }
    # set_column_widths(ws_detailed, detailed_min_widths)
    
    # # 3. Line-wise Comparison Sheet
    # ws_comparison = wb.create_sheet("Line Comparison")
    
    # # Create comparison data
    # comparison_data = []
    # sections = ['Assembly', 'Back', 'Collar', 'Cuff', 'Front', 'Sleeve']
    
    # for line, line_data in data.items():
    #     row = {'Line': line}
    #     for section in sections:
    #         predicted = line_data['predicted_absenteeism_percentage'].get(section, 0)
    #         actual = line_data['actual_absenteeism_percentage'].get(section, 0)
    #         row[f'{section}_Predicted%'] = predicted
    #         row[f'{section}_Actual%'] = actual
    #         row[f'{section}_Variance'] = round(actual - predicted, 2)
    #     comparison_data.append(row)
    
    # df_comparison = pd.DataFrame(comparison_data)
    
    # # Write comparison data to sheet
    # for r in dataframe_to_rows(df_comparison, index=False, header=True):
    #     ws_comparison.append(r)
    
    # # Format comparison sheet
    # for cell in ws_comparison[1]:
    #     cell.font = header_font
    #     cell.fill = header_fill
    #     cell.alignment = Alignment(horizontal='center')
    #     cell.border = border
    
    # # Set column widths with specific minimums for comparison sheet
    # comparison_min_widths = {
    #     'Line': 8,
    #     # Predicted columns
    #     'Assembly_Predicted%': 16,
    #     'Back_Predicted%': 14,
    #     'Collar_Predicted%': 15,
    #     'Cuff_Predicted%': 14,
    #     'Front_Predicted%': 15,
    #     'Sleeve_Predicted%': 16,
    #     # Actual columns
    #     'Assembly_Actual%': 14,
    #     'Back_Actual%': 12,
    #     'Collar_Actual%': 13,
    #     'Cuff_Actual%': 12,
    #     'Front_Actual%': 13,
    #     'Sleeve_Actual%': 14,
    #     # Variance columns
    #     'Assembly_Variance': 16,
    #     'Back_Variance': 13,
    #     'Collar_Variance': 14,
    #     'Cuff_Variance': 13,
    #     'Front_Variance': 14,
    #     'Sleeve_Variance': 15
    # }
    # set_column_widths(ws_comparison, comparison_min_widths)
    
    # # 4. Accuracy Analysis Sheet
    # ws_accuracy = wb.create_sheet("Accuracy Analysis")
    
    # accuracy_data = []
    # for line, line_data in data.items():
    #     for section in line_data['total_employee_count'].keys():
    #         predicted_pct = line_data['predicted_absenteeism_percentage'].get(section, 0)
    #         actual_pct = line_data['actual_absenteeism_percentage'].get(section, 0)
            
    #         # Calculate different accuracy metrics
    #         absolute_error = abs(actual_pct - predicted_pct)
    #         percentage_error = (absolute_error / max(actual_pct, 1)) * 100 if actual_pct > 0 else 0
            
    #         accuracy_data.append({
    #             'Line': line,
    #             'Section': section,
    #             'Predicted %': predicted_pct,
    #             'Actual %': actual_pct,
    #             'Absolute Error': round(absolute_error, 2),
    #             'Percentage Error': round(percentage_error, 2),
    #             'Accuracy Score': round(100 - absolute_error, 2),
    #             'Direction': 'Over-predicted' if predicted_pct > actual_pct else 'Under-predicted' if predicted_pct < actual_pct else 'Exact'
    #         })
    
    # df_accuracy = pd.DataFrame(accuracy_data)
    
    # # Write accuracy data to sheet
    # for r in dataframe_to_rows(df_accuracy, index=False, header=True):
    #     ws_accuracy.append(r)
    
    # # Format accuracy sheet
    # for cell in ws_accuracy[1]:
    #     cell.font = header_font
    #     cell.fill = header_fill
    #     cell.alignment = Alignment(horizontal='center')
    #     cell.border = border
    
    # # Set column widths with specific minimums for accuracy sheet
    # accuracy_min_widths = {
    #     'Line': 8,
    #     'Section': 10,
    #     'Predicted %': 12,
    #     'Actual %': 10,
    #     'Absolute Error': 14,
    #     'Percentage Error': 16,
    #     'Accuracy Score': 14,
    #     'Direction': 15
    # }
    # set_column_widths(ws_accuracy, accuracy_min_widths)

    # Save Excel file in memory using BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)  # Move to the beginning of the BytesIO stream

    return output  # Returning the Excel file as a byte stream




def write_absenteeism_data_to_csv(data, filename="summary.csv"):
    """
    Generates a summary CSV from absenteeism data and returns it as a BytesIO stream.
    
    Args:
        data: Dictionary containing absenteeism data for different production lines
        filename: Name to use for the CSV file (for email attachment)
    
    Returns:
        Tuple of (BytesIO object, filename)
    """
    if isinstance(data, str):
        data = json.loads(data)

    # Remove 'prediction_date' if it exists
    data.pop('prediction_date', None)
    # Remove 'generated_at' if it exists
    data.pop('generated_at', None)

    summary_data = []
    for line, line_data in data.items():
        total_employees = sum(line_data['total_employee_count'].values())
        total_predicted = sum([item['count'] for item in line_data['predicted_absent_count']])
        total_actual = sum([item['count'] for item in line_data['actual_absent_count']])
        predicted_avg = line_data.get('predicted_absenteeism_percentage', 0)
        actual_avg = line_data.get('actual_absenteeism_percentage', 0)

        summary_data.append({
            'Line': line.title(),
            'Total Employees': total_employees,
            'Predicted Absent': int(total_predicted),
            'Actual Absent': total_actual,
            'Predicted Absenteeism %': round(predicted_avg, 2),
            'Actual Absenteeism %': round(actual_avg, 2)
        })

    df_summary = pd.DataFrame(summary_data)

    # Write to in-memory buffer
    buffer = BytesIO()
    df_summary.to_csv(buffer, index=False)
    buffer.seek(0)
    
    return buffer, filename



def export_absenteeism_predictions_excel():
    # Fetch required fields only (avoid extra columns)
    qs = AbsenteeismPrediction.objects.values(
        "datetime",
        "day_of_week",
        "predicted_absent_count",
        "line",
        "section",
        "forecast_period",
    )

    # Convert queryset → DataFrame
    df = pd.DataFrame.from_records(qs)

    # Rename columns for clarity (optional)
    df.rename(columns={
        "datetime": "Date",
        "day_of_week": "Day of Week",
        "predicted_absent_count": "Predicted Absent Count",
        "line": "Line",
        "section": "Section",
        "forecast_period": "Forecast Period",
    }, inplace=True)

    # Create response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="Prediction_Data_2_Months.xlsx"'

    # Write Excel file
    with pd.ExcelWriter(response, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Absenteeism Predictions", index=False)

        # Access the worksheet
        worksheet = writer.sheets["Absenteeism Predictions"]

        # Adjust column widths
        for col_idx, col in enumerate(df.columns, 1):  # 1-based index for openpyxl
            max_length = max(
                df[col].astype(str).map(len).max(),
                len(col)  # include header length
            ) + 2  # padding
            worksheet.column_dimensions[get_column_letter(col_idx)].width = max_length

    return response
